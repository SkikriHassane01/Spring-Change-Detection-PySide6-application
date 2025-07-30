from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFrame,
    QTabWidget, QTableView, QHeaderView, QFileDialog, QMessageBox,
    QScrollArea, QSizePolicy, QApplication
)
from PySide6.QtCore import Qt, QAbstractTableModel, QModelIndex
from PySide6.QtGui import QFont, QColor, QBrush

import pandas as pd
import os
import io
from openpyxl import load_workbook

from utils.app_state import AppState
from utils.file_handler import FileHandler
from gui.components.excel_table_model import DataFrameModel
from gui.styling import AppStyles

class ColoredDataFrameModel(DataFrameModel):
    """Extended DataFrameModel with row coloring based on change type."""
    
    def __init__(self, data=None):
        super().__init__(data)
        # Map of Change Type to colors
        self.color_map = {
            "New": QColor("#FF5733"),        # Red for new cars
            "Spring Changed": QColor("#CDA000"),  # Custom yellow for changed cars
        }
    
    def data(self, index, role=Qt.DisplayRole):
        """Get data at the given index with additional styling."""
        if role == Qt.BackgroundRole:
            # Apply highlighting based on Change Type column
            if index.isValid() and "Change Type" in self.dataframe.columns:
                row = index.row()
                change_type = self.dataframe.iloc[row]["Change Type"]
                
                if change_type in self.color_map:
                    return self.color_map[change_type]
            
        elif role == Qt.ForegroundRole:
            # Make text white for red background for readability
            if index.isValid() and "Change Type" in self.dataframe.columns:
                row = index.row()
                change_type = self.dataframe.iloc[row]["Change Type"]
                
                if change_type == "New":
                    return QBrush(QColor("white"))
        
        return super().data(index, role)


class ResultsWidget(QWidget):
    """Widget for displaying analysis results."""
    
    def __init__(self, state: AppState, parent=None):
        super().__init__(parent)
        self.state = state
        self.image_max_width = 800  # Maximum width for displayed images
        self.init_ui()
        
        # Update results when the widget is created
        if self.state.results_df is not None:
            self.update_results()
    
    def showEvent(self, event):
        """Handle widget becoming visible."""
        super().showEvent(event)
        # Update results when the widget becomes visible
        if self.state.results_df is not None:
            self.update_results()
    
    def create_moteur_list_tab(self):
        """Create a tab showing the list of unique motors (Moteur column)."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        moteur_list_label = QLabel("🔧 Unique Motors List")
        moteur_list_label.setFont(QFont("Arial", 14, QFont.Bold))
        moteur_list_label.setStyleSheet("color: white; margin-bottom: 10px;")
        layout.addWidget(moteur_list_label)
        
        if self.state.new_df is None or "Moteur" not in self.state.new_df.columns:
            no_data_label = QLabel("No 'Moteur' column found in the uploaded data.")
            no_data_label.setStyleSheet("color: #FFAA00; font-style: italic;")
            layout.addWidget(no_data_label)
            return tab

        # Get sorted unique motors
        moteurs = sorted(self.state.new_df["Moteur"].dropna().unique())

        # Display as a scrollable label list
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)

        for moteur in moteurs:
            moteur_label = QLabel(f"• {moteur}")
            moteur_label.setStyleSheet("color: #CCCCCC; font-size: 13px; padding: 2px;")
            scroll_layout.addWidget(moteur_label)

        scroll_area.setWidget(scroll_content)
        layout.addWidget(scroll_area)

        return tab

    def init_ui(self):
        """Initialize the user interface."""
        # Apply dark mode style
        self.setStyleSheet(AppStyles.DARK_MODE_STYLE)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Header
        header_label = QLabel("📊 Step 3: Analysis Results")
        header_label.setFont(QFont("Arial", 16, QFont.Bold))
        header_label.setStyleSheet(f"color: white; background-color: {AppStyles.PRIMARY_COLOR}; padding: 8px; border-radius: 4px;")
        layout.addWidget(header_label)
        
        # Description
        description = QLabel("Here are your comprehensive analysis results:")
        description.setStyleSheet(AppStyles.LABEL_STYLE)
        layout.addWidget(description)
        
        # Tabs for results
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet(f"""
            QTabWidget::pane {{
                border: 1px solid {AppStyles.DARK_BORDER};
                border-radius: 4px;
                background-color: {AppStyles.DARK_CARD_BG};
            }}
            QTabBar::tab {{
                background-color: {AppStyles.DARK_HIGHLIGHT};
                color: {AppStyles.TEXT_COLOR};
                border: 1px solid {AppStyles.DARK_BORDER};
                border-bottom: none;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                padding: 8px 12px;
                margin-right: 2px;
            }}
            QTabBar::tab:selected {{
                background-color: #003366;
                color: white;
                border-color: #002244;
                font-weight: bold;
            }}
            QTabBar::tab:hover:!selected {{
                background-color: #3D3D3D;
            }}
        """)
        
        layout.addWidget(self.tabs)
        
        # Download button
        button_layout = QHBoxLayout()
        
        self.download_button = QPushButton("📄 Download Excel Report")
        self.download_button.setMinimumHeight(40)
        self.download_button.setStyleSheet(AppStyles.ACTIVE_BUTTON_STYLE)
        self.download_button.clicked.connect(self.download_results)
        button_layout.addWidget(self.download_button)
        
        # Reset button
        self.reset_button = QPushButton("📁 Upload New Files")
        self.reset_button.setMinimumHeight(40)
        self.reset_button.setStyleSheet(AppStyles.INACTIVE_BUTTON_STYLE)
        self.reset_button.clicked.connect(self.reset_workflow)
        button_layout.addWidget(self.reset_button)
        
        layout.addLayout(button_layout)
        
        # Legend
        legend_frame = QFrame()
        legend_frame.setFrameShape(QFrame.StyledPanel)
        legend_frame.setStyleSheet(f"background-color: {AppStyles.DARK_CARD_BG}; border: 1px solid {AppStyles.DARK_BORDER}; border-radius: 4px; padding: 10px;")
        legend_layout = QVBoxLayout(legend_frame)
        
        legend_title = QLabel("Color Legend:")
        legend_title.setFont(QFont("Arial", 10, QFont.Bold))
        legend_title.setStyleSheet("color: white;")
        legend_layout.addWidget(legend_title)
        
        new_legend = QLabel("🟥 New rows: Cars added in new PTA file")
        new_legend.setStyleSheet("color: #FF5733; font-weight: bold;")
        legend_layout.addWidget(new_legend)
        
        changed_legend = QLabel("🟨 Spring Changed: Reference (spring) changed")
        changed_legend.setStyleSheet("color: #CDA000; font-weight: bold;")
        legend_layout.addWidget(changed_legend)
        
        layout.addWidget(legend_frame)
        
        # Initialize tabs
        self.update_results()
    
    def update_results(self):
        """Update the UI with results data following Streamlit tab structure."""
        # Clear all existing tabs
        self.tabs.clear()
        
        if self.state.results_df is None:
            # No results available
            no_results_tab = QWidget()
            no_results_layout = QVBoxLayout(no_results_tab)
            
            message = QLabel("No results available. Please run the analysis first.")
            message.setStyleSheet("color: #FFC107; font-weight: bold;")
            no_results_layout.addWidget(message)
            
            self.tabs.addTab(no_results_tab, "Analysis Results")
            return
        
        # Load Excel data if available and not already loaded
        if self.state.new_file_path and not self.state.excel_sheets_data:
            self.load_excel_data()
        
        # Tab order: Analysis Results, Data Sheets, Assiette Graphs
        
        # Tab 1: Analysis Results
        results_tab = self.create_analysis_results_tab()
        self.tabs.addTab(results_tab, "Analysis Results")
        
        # Find Assiette théorique sheet first
        last_sheet = None
        for sheet_name in self.state.excel_sheets_data.keys():
            if any(s in sheet_name.lower() for s in ["assiette", "théorique", "theorique"]):
                last_sheet = sheet_name
                break
        
        # If found, add its graphs tab right after results
        if last_sheet:
            last_sheet_graphs = f"{last_sheet} Graphs"
            if last_sheet_graphs in self.state.excel_graphs_data:
                graphs_tab = self.create_sheet_graphs_tab(last_sheet, self.state.excel_graphs_data[last_sheet_graphs])
                self.tabs.addTab(graphs_tab, f"{last_sheet} Graphs")
        
        # Tab 2: Entete (if available)
        if "Entete" in self.state.excel_sheets_data:
            entete_tab = self.create_sheet_tab("Entete", self.state.excel_sheets_data["Entete"])
            self.tabs.addTab(entete_tab, "Entete")
        
        # Tab 3: Options (if available)
        if "Options" in self.state.excel_sheets_data:
            options_tab = self.create_sheet_tab("Options", self.state.excel_sheets_data["Options"])
            self.tabs.addTab(options_tab, "Options")
        
        # Tab 4: Assiette théorique (if available)
        if last_sheet:
            sheet_tab = self.create_sheet_tab(last_sheet, self.state.excel_sheets_data[last_sheet])
            self.tabs.addTab(sheet_tab, "Assiette Théorique")
    
        # Tab 5: Moteur List (if Moteur column is available)
        if self.state.new_df is not None and "Moteur" in self.state.new_df.columns:
            moteur_tab = self.create_moteur_list_tab()
            self.tabs.addTab(moteur_tab, "Moteur List")
    def load_excel_data(self):
        """Load Excel sheet data in background."""
        if not self.state.new_file_path or not os.path.exists(self.state.new_file_path):
            return
            
        try:
            QApplication.setOverrideCursor(Qt.WaitCursor)
            sheets_data, graphs_data = FileHandler.extract_sheets_and_graphs(self.state.new_file_path)
            self.state.excel_sheets_data = sheets_data
            self.state.excel_graphs_data = graphs_data
            QApplication.restoreOverrideCursor()
        except Exception as e:
            QApplication.restoreOverrideCursor()
            print(f"Could not extract Excel data: {str(e)}")
    
    def create_analysis_results_tab(self):
        """Create the analysis results tab with highlighting."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Prepare display data
        display_df = self.prepare_display_data()
        
        # Create table view with colored model - more compact
        table_view = QTableView()
        table_view.setMinimumHeight(250)
        table_view.setMaximumHeight(400)
        table_view.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        table_view.setStyleSheet(AppStyles.TABLE_STYLE)
        
        # Set up the model with display data
        model = ColoredDataFrameModel(display_df)
        table_view.setModel(model)
        
        # Configure column widths
        table_view.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        
        # Set reasonable default widths
        for col in range(model.columnCount()):
            if col < len(model.dataframe.columns):
                col_name = model.dataframe.columns[col]
                
                if 'reference' in col_name.lower() or 'référence' in col_name.lower():
                    table_view.setColumnWidth(col, 180)
                elif 'masse' in col_name.lower() or 'mass' in col_name.lower():
                    table_view.setColumnWidth(col, 100)
                elif 'change' in col_name.lower() or 'status' in col_name.lower():
                    table_view.setColumnWidth(col, 150)
                else:
                    table_view.setColumnWidth(col, 120)
        
        table_view.setSortingEnabled(True)
        layout.addWidget(table_view)
        
        return tab
    
    def prepare_display_data(self):
        """Prepare display data by merging new data with results metadata."""
        if self.state.new_df is None or self.state.results_df is None:
            return self.state.results_df if self.state.results_df is not None else pd.DataFrame()
        
        # Create display dataframe starting with new data
        display_df = self.state.new_df.copy()
        
        # Add metadata columns from results
        metadata_cols = [
            'Old Reference', 'New Reference',
            'Mass Status', 'Change Type',
            'Cell ID New', 'Cell ID Old'
        ]
        
        # Add metadata columns that exist in results
        for col in metadata_cols:
            if col in self.state.results_df.columns:
                if len(display_df) == len(self.state.results_df):
                    display_df[col] = self.state.results_df[col].values
        
        # Sort by Cell ID New for consistent display
        if 'Cell ID New' in display_df.columns:
            display_df = display_df.sort_values('Cell ID New', ascending=True).reset_index(drop=True)
        
        return display_df
    

    def add_other_sheets_tabs(self):
        """Add tabs for other sheets and their graphs."""
        # Get sheet names excluding PTA
        sheet_names = [name for name in self.state.excel_sheets_data.keys() if name.upper() != "PTA"]
        
        # Sort sheets - put "Assiette théorique" at the end
        regular_sheets = []
        special_sheets = []
        
        for sheet_name in sheet_names:
            if any(special in sheet_name.lower() for special in ["assiette", "théorique", "theorique"]):
                special_sheets.append(sheet_name)
            else:
                regular_sheets.append(sheet_name)
        
        # Add regular sheets first, then special sheets
        ordered_sheets = regular_sheets + special_sheets
        
        for sheet_name in ordered_sheets:
            # Create sheet data tab
            sheet_tab = self.create_sheet_tab(sheet_name, self.state.excel_sheets_data[sheet_name])
            self.tabs.addTab(sheet_tab, sheet_name)
            
            # Add graphs tab if available
            graph_key = f"{sheet_name} Graphs"
            if graph_key in self.state.excel_graphs_data:
                graphs_tab = self.create_sheet_graphs_tab(sheet_name, self.state.excel_graphs_data[graph_key])
                
                # Special naming for Assiette théorique
                is_special = any(s in sheet_name.lower() for s in ["assiette", "théorique", "theorique"])
                tab_title = "Assiette Théorique" if is_special else f"{sheet_name} Charts"
                
                self.tabs.addTab(graphs_tab, tab_title)
    
    def create_sheet_tab(self, sheet_name, sheet_data):
        """Create a tab for displaying sheet data."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        if isinstance(sheet_data, pd.DataFrame):
            # Create table view for sheet data - more compact
            table_view = QTableView()
            table_view.setMinimumHeight(250)
            table_view.setMaximumHeight(400)
            table_view.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            table_view.setStyleSheet(AppStyles.TABLE_STYLE)
            
            # Set up the model
            model = DataFrameModel(sheet_data)
            table_view.setModel(model)
            
            # Configure table
            table_view.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            table_view.setSortingEnabled(True)
            
            layout.addWidget(table_view)
        else:
            # Display error message
            error_label = QLabel(str(sheet_data))
            error_label.setStyleSheet("color: #FF5733;")
            layout.addWidget(error_label)
        
        # Add sheet name caption
        caption = QLabel(f"Sheet: {sheet_name}")
        caption.setStyleSheet("color: #666666; font-style: italic; margin-top: 10px;")
        layout.addWidget(caption)
        
        return tab
    
    def create_sheet_graphs_tab(self, sheet_name, graphs_data):
        """Create a tab for sheet graphs with special handling."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        if not graphs_data:
            label = QLabel(f"No charts found for {sheet_name}")
            label.setStyleSheet("color: #666666;")
            layout.addWidget(label)
            return tab
        
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        
        # Special handling for Assiette théorique
        is_special = any(s in sheet_name.lower() for s in ["assiette", "théorique", "theorique"])
        title = "Assiette Théorique" if is_special else f"{sheet_name} Graphs"
        
        title_label = QLabel(title)
        title_label.setFont(QFont("Arial", 14, QFont.Bold))
        title_label.setStyleSheet("color: white; margin-bottom: 10px;")
        scroll_layout.addWidget(title_label)
        
        # Display graphs with special sizing for important sheets
        self.display_graphs(scroll_layout, graphs_data, is_special=is_special)
        
        scroll_area.setWidget(scroll_content)
        layout.addWidget(scroll_area)
        
        return tab
    
    def display_graphs(self, layout, graphs_data, is_special=False):
        """Display graphs with appropriate sizing."""
        for i, graph in enumerate(graphs_data):
            if isinstance(graph, dict) and 'data' in graph:
                try:
                    # Convert base64 to QPixmap
                    from PySide6.QtGui import QPixmap
                    from PySide6.QtCore import QByteArray
                    import base64
                    
                    img_data = base64.b64decode(graph['data'])
                    pixmap = QPixmap()
                    pixmap.loadFromData(QByteArray(img_data))
                    
                    # Calculate display size
                    if 'width' in graph and 'height' in graph:
                        orig_width = graph['width']
                        orig_height = graph['height']
                        
                        # Special sizing for important graphs
                        max_width = int(self.image_max_width * 1.2) if is_special else self.image_max_width
                        display_width = min(max_width, orig_width)
                        
                        if orig_width > 0:
                            scale_factor = display_width / orig_width
                            display_height = int(orig_height * scale_factor)
                            pixmap = pixmap.scaled(display_width, display_height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    
                    # Create label and set pixmap
                    img_label = QLabel()
                    img_label.setPixmap(pixmap)
                    img_label.setAlignment(Qt.AlignCenter)
                    img_label.setFrameShape(QFrame.Box)
                    img_label.setFrameShadow(QFrame.Sunken)
                    img_label.setLineWidth(1)
                    img_label.setStyleSheet("border: 1px solid #DDDDDD; background-color: white; margin: 10px;")
                    
                    layout.addWidget(img_label)
                    
                    # Add separator between graphs
                    if i < len(graphs_data) - 1:
                        line = QFrame()
                        line.setFrameShape(QFrame.HLine)
                        line.setFrameShadow(QFrame.Sunken)
                        layout.addWidget(line)
                        
                except Exception as e:
                    error_label = QLabel(f"Error displaying chart {i+1}: {str(e)}")
                    error_label.setStyleSheet("color: #FF5733;")
                    layout.addWidget(error_label)
    
    def download_results(self):
        """Download results as Excel file with highlighting."""
        if self.state.results_df is None:
            QMessageBox.warning(self, "No Results", "No analysis results available to download.")
            return
        
        try:
            QApplication.setOverrideCursor(Qt.WaitCursor)
            
            # Get save file path
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Save Results",
                os.path.expanduser("~/spring_change_analysis.xlsx"),
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                QApplication.restoreOverrideCursor()
                return
                
            # Add .xlsx extension if missing
            if not file_path.endswith('.xlsx'):
                file_path += '.xlsx'
            
            # Create Excel file with highlighting
            if self.state.new_file_path and os.path.exists(self.state.new_file_path):
                try:
                    # Use the original file as template
                    excel_bytes = FileHandler.create_excel_bytes(
                        self.state.results_df,
                        self.state.new_file_path
                    )
                    
                    with open(file_path, 'wb') as f:
                        f.write(excel_bytes)
                    
                    # Store the path of saved report
                    self.last_saved_report = file_path
                    
                    # Directly open the file without confirmation
                    if os.name == 'nt':  # Windows
                        os.startfile(file_path)
                    elif os.name == 'posix':  # macOS and Linux
                        from subprocess import call
                        call(('open' if os.name == 'darwin' else 'xdg-open', file_path))
                except Exception as e:
                    QApplication.restoreOverrideCursor()
                    # Fall back to simple export
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        self.state.results_df.to_excel(writer, sheet_name="Analysis Results", index=False)
                    
                    QMessageBox.information(
                        self, 
                        "Export Complete", 
                        f"Results saved successfully to:\n{file_path}\n\n(Simple format - highlighting failed: {str(e)})"
                    )
            else:
                # Basic export
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    self.state.results_df.to_excel(writer, sheet_name="Analysis Results", index=False)
                
                QMessageBox.information(
                    self, 
                    "Export Complete", 
                    f"Results saved successfully to:\n{file_path}"
                )
                
            QApplication.restoreOverrideCursor()
                
        except Exception as e:
            QApplication.restoreOverrideCursor()
            QMessageBox.critical(
                self,
                "Export Error",
                f"Error saving results: {str(e)}"
            )
    
    def reset_workflow(self):
        """Reset the application to the upload step and clear previous files."""
        # Confirm reset
        answer = QMessageBox.question(
            self,
            "Reset Workflow",
            "Are you sure you want to restart and upload new files? All current analysis will be lost.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if answer != QMessageBox.Yes:
            return

        # 1) Clear AppState
        self.state.reset_data()

        # 2) Reset upload widget UI using the main window reference
        main_win = self.window()
        uw = getattr(main_win, 'upload_widget', None)
        if uw:
            # Switch back to Old PTA File tab
            uw.tabs.setCurrentIndex(0)
            # Clear both upload frames
            for frame in (uw.old_file_frame, uw.new_file_frame):
                frame.df = None
                frame._last_file_path = None
                frame.status_label.setText("No file selected")
                frame.status_label.setStyleSheet(
                    "color: #888888; font-style: italic; font-size: 11px;"
                )
                frame.table_view.setModel(None)
                frame.table_view.setVisible(False)
            # Hide status and proceed controls
            uw.update_status()

        # 3) Navigate back to upload step on main window
        if hasattr(main_win, 'navigate_to_step'):
            main_win.navigate_to_step(0)
        # Ensure sidebar button highlight
        if hasattr(main_win, 'update_workflow_buttons'):
            main_win.update_workflow_buttons(0)
