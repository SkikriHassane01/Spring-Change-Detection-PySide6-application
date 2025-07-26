"""
This module handles file operations for the Spring Change Detection application.
It validates Excel files, processes data, and creates output files.
"""
from typing import Tuple, Optional, Any, Dict
import os
import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import base64

# Configuration - similar to original application
UPLOAD_CONFIG = {
    "allowed_extension": ['xlsx', 'xls'],
    "max_file_size": 200,  # MB
    "sheet_name": "PTA",
    "skip_rows": [1]
}

REQUIRED_COLUMNS = {
    "mass": "Masse suspendue en charge de référence",
    "reference": "Référence"
}

class FileHandler:
    """Handles validation and export of Excel files."""

    @staticmethod
    def validate_excel_file(file_path: str, file_label: str) -> Tuple[bool, str, Optional[pd.DataFrame]]:
        """
        Validate an Excel file at the given path.
        
        Args:
            file_path: Path to the Excel file.
            file_label: A label for the file (e.g., "old", "new").
            
        Returns:
            Tuple containing:
              - validity (bool)
              - message (str)
              - DataFrame if valid, else None
        """
        if not file_path:
            return False, f"No '{file_label}' file selected.", None
        
        if not os.path.exists(file_path):
            return False, f"File '{file_path}' does not exist.", None
        
        file_ext = os.path.splitext(file_path)[1][1:].lower()
        if file_ext not in UPLOAD_CONFIG["allowed_extension"]:
            return False, f"Invalid file format. Please upload an Excel file (.xlsx or .xls).", None
        
        file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
        if file_size_mb > UPLOAD_CONFIG["max_file_size"]:
            return False, f"File size exceeds {UPLOAD_CONFIG['max_file_size']} MB.", None
        
        try:
            df = pd.read_excel(
                file_path,
                engine="openpyxl",
                sheet_name=UPLOAD_CONFIG["sheet_name"],
                skiprows=UPLOAD_CONFIG["skip_rows"],
            ).reset_index(drop=True)
        except Exception as e:
            return False, f"Error reading '{file_label}' file: {e}", None
        
        if df.empty:
            return False, f"'{file_label}' file is empty.", None
        
        is_valid, msg = FileHandler._validate_columns(df)
        if not is_valid:
            return False, msg, None
        
        return True, "File uploaded successfully.", df
    
    @staticmethod
    def _validate_columns(df: pd.DataFrame) -> Tuple[bool, str]:
        """
        Check for required columns in the DataFrame.
        
        Args:
            df: DataFrame to validate.
            
        Returns:
            Tuple with validity status and error message.
        """
        required_cols = [
            REQUIRED_COLUMNS["mass"],
            REQUIRED_COLUMNS["reference"],
        ]
        missing = [col for col in required_cols if col not in df.columns]
        if missing:
            return False, f"Missing columns: {', '.join(missing)}."
        return True, ""
    
    @staticmethod
    def create_excel_bytes(results_df: pd.DataFrame, original_file_path: str = None) -> bytes:
        """
        Generate Excel report that preserves the original file structure 
        and only adds highlighting to changed/new rows.
        
        Args:
            results_df: Analysis results with metadata.
            original_file_path: Path to the original Excel file.
            
        Returns:
            Byte content of the Excel file with highlighting.
        """
        output = io.BytesIO()
        
        if not original_file_path or not os.path.exists(original_file_path):
            raise ValueError("Original file path is required and must exist.")
        
        try:
            # Load the original workbook
            wb = load_workbook(original_file_path)
            
            # Get the PTA sheet
            if UPLOAD_CONFIG["sheet_name"] in wb.sheetnames:
                ws = wb[UPLOAD_CONFIG["sheet_name"]]
                
                # Find data start row (after skipping header rows)
                start_row = UPLOAD_CONFIG["skip_rows"][0] + 2  # Skip header + extra row
                
                # Create a mapping of Cell ID to Change Type for faster lookup
                cell_id_to_change = {}
                if 'Cell ID New' in results_df.columns and 'Change Type' in results_df.columns:
                    for _, row in results_df.iterrows():
                        cell_id_to_change[row['Cell ID New']] = row['Change Type']
                
                # Apply highlighting to rows based on analysis results
                row_idx = start_row
                max_iterations = 10000  # Prevent infinite loops
                iterations = 0
                
                while iterations < max_iterations:
                    # Check if we've reached the end of data
                    if ws.cell(row=row_idx, column=1).value is None:
                        break
                    
                    # Get cell ID (Excel row number)
                    cell_id = row_idx
                    
                    # Check if this row has changes
                    if cell_id in cell_id_to_change:
                        change_type = cell_id_to_change[cell_id]
                        
                        if change_type == 'New':
                            # Highlight new rows with red background and white text
                            fill = PatternFill('solid', fgColor='FF5733')
                            font = Font(color='FFFFFF')  # White text for better contrast
                            
                            for col in range(1, ws.max_column + 1):
                                cell = ws.cell(row=row_idx, column=col)
                                cell.fill = fill
                                cell.font = font
                                
                        elif change_type == 'Spring Changed':
                            # Highlight spring changed rows with blue background
                            fill = PatternFill('solid', fgColor='B4C6E7')
                            font = Font(color='000000')  # Black text for better contrast
                            
                            for col in range(1, ws.max_column + 1):
                                cell = ws.cell(row=row_idx, column=col)
                                cell.fill = fill
                                cell.font = font
                    
                    row_idx += 1
                    iterations += 1
            
            # Save the modified workbook
            wb.save(output)
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            # If highlighting fails, return original file
            print(f"Error adding highlighting: {str(e)}")
            with open(original_file_path, 'rb') as f:
                return f.read()
    
    @staticmethod
    def extract_sheets_and_graphs(file_path: str) -> Tuple[Dict[str, pd.DataFrame], Dict[str, Any]]:
        """
        Extract all sheets and graphical elements from an Excel file.
        
        Args:
            file_path: Path to the Excel file.
            
        Returns:
            Tuple with dictionaries of sheets data and graphs data.
        """
        sheets_data = {}
        graphs_data = {}
        
        if not os.path.exists(file_path):
            return sheets_data, graphs_data
        
        try:
            # Get all sheet names
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            
            # Extract sheet data
            for sheet_name in sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    sheets_data[sheet_name] = df
                except Exception as e:
                    sheets_data[sheet_name] = f"Error loading sheet: {str(e)}"
            
            # Extract charts/images
            try:
                wb = load_workbook(file_path, data_only=True)
                
                # Process each sheet
                for sheet_name in wb.sheetnames:
                    sheet_graphs = []
                    ws = wb[sheet_name]
                    
                    # Extract images
                    if hasattr(ws, '_images') and ws._images:
                        for img in ws._images:
                            try:
                                img_data = img._data()
                                if img_data:
                                    # Convert to base64 for storing
                                    img_b64 = base64.b64encode(img_data).decode()
                                    
                                    # Store with metadata
                                    sheet_graphs.append({
                                        'data': img_b64,
                                        'width': img.width,
                                        'height': img.height
                                    })
                            except Exception:
                                continue
                    
                    # Store graphs for this sheet
                    if sheet_graphs:
                        graphs_data[f"{sheet_name} Graphs"] = sheet_graphs
            except Exception as e:
                # Log error but continue
                print(f"Error extracting graphs: {str(e)}")
            
        except Exception as e:
            # Log error but return what we have
            print(f"Error extracting sheets: {str(e)}")
        
        return sheets_data, graphs_data